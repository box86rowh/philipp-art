from django.shortcuts import render_to_response
from django.shortcuts import render
from django.core.files.uploadedfile import UploadedFile
from django.template import RequestContext
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.core.urlresolvers import reverse
from django.contrib import auth
from django.http import HttpResponseRedirect
from django.http import HttpResponse
from .models import Location
from .models import ArtPiece
from .models import ArtPiecePhoto
from django.shortcuts import get_object_or_404
from sorl.thumbnail import get_thumbnail
import logging
import json
import openpyxl
from openpyxl.cell import get_column_letter
from openpyxl.styles import Alignment, Style, Font
from openpyxl.drawing import Image
from philipp_art.settings import PROJECT_ROOT

@login_required
def home(request):
    locations = Location.objects.all()
    return render_to_response('home.html',
                          {"locations": locations},
                          context_instance=RequestContext(request))
@login_required
def art_list(request,id=None):
    if id == None:
        location = None
        pieces = ArtPiece.objects.all()
    else:
        location = Location.objects.get(pk = id)
        pieces = location.pieces()

    return render_to_response('art_list.html',
                              {"pieces": pieces, "location" : location},
                              context_instance=RequestContext(request))

def login(request):
    if request.method == 'POST' :
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = auth.authenticate(username = username, password = password)      
    
        if user is not None:
            auth.login(request, user)
            return HttpResponseRedirect(reverse('home'))
        else:
            return HttpResponseRedirect('/accounts/login')
    else:
        return render(request, 'login.html')
    
@login_required
def edit_location(request, id):
    if id != '0' :
        location = get_object_or_404(Location,pk=id)
    else:
        location = Location()
    if request.method == 'POST' :
        location.title = request.POST.get('title', '')
        location.address = request.POST.get('address', '')
        location.save()
        return HttpResponseRedirect('/locations/%s/edit' % location.pk)
    else:
        return render_to_response('edit_location.html',
                              {"location" : location},
                              context_instance=RequestContext(request))
@login_required
def edit_piece(request, id):
    locations = Location.objects.all()
    if id != '0' :
        piece = get_object_or_404(ArtPiece,pk=id)
    else:
        piece = ArtPiece()
    if request.method == 'POST' :
        piece.title = request.POST.get('title', '')
        piece.artist = request.POST.get('artist', '')
        piece.notes = request.POST.get('notes', '')
        piece.location = Location.objects.get(pk = request.POST.get('location', ''))
        piece.purchase_date = request.POST.get('purchase_date', '')
        piece.purchase_price = request.POST.get('purchase_price', '')
        piece.save()
        return HttpResponseRedirect('/pieces/%s/edit' % piece.pk)
    else:
        return render_to_response('edit_piece.html',
                                  {"piece" : piece, "locations": locations,},
                                  context_instance=RequestContext(request))
    
@login_required
@csrf_exempt
def delete_piece(request, piece_id):
    piece = get_object_or_404(ArtPiece,pk=piece_id)
    piece.delete()
    return HttpResponse("OK")


@login_required
@csrf_exempt
def delete_photo(request, piece_id, photo_id):
    image = ArtPiecePhoto.objects.get(pk=photo_id)
    image.delete()
    return HttpResponse("OK")

@login_required
@csrf_exempt
def export(request):
    clist = request.GET.get('include', '')
    joe = clist.split(',')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "MyModel"

    row_num = 0

    columns = [
        (u"ID", 5),
        (u"Title", 40),
        (u"Artist", 40),
        (u"Purchase Date", 20),
        (u"Purchase Price", 20),
        (u"Notes", 40),
        (u"Image", 60),
    ]
    
    alignment = Alignment(wrap_text=True)
    
    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style = Style(font=Font(bold=True))
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for piece in ArtPiece.objects.filter(pk__in=joe):
        row_num += 1
        row = [
            piece.pk,
            piece.title,
            piece.artist,
            piece.purchase_date,
            piece.purchase_price,
            piece.notes,
        ]
        
        for col_num in xrange(len(row)):
            print col_num + 1
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style = c.style.copy(alignment = alignment)
        
        # url = PROJECT_ROOT + piece.photo().url
        # thumb = get_thumbnail(url, '100x100', crop='center', quality=99)
        # img = Image(PROJECT_ROOT + thumb.url)        
        # c = ws.cell(row=row_num + 1, column=20)
        # img.anchor(c)
        # ws.add_image(img)

    wb.save(response)
    return response    
    

@login_required
@csrf_exempt
def add_photo_to_piece(request, id):
    log = logging.getLogger(__name__)
    if request.method == 'POST':
        print 'POST'
        if request.FILES == None:
            return HttpResponseBadRequest('Must have files attached!')
        #getting file data for farther manipulations
        file = request.FILES[u'files[]']
        wrapped_file = UploadedFile(file)
        filename = wrapped_file.name
        file_size = wrapped_file.file.size
        print 'Got file: "'+str(filename)+'"'
    
        #writing file manually into model
        #because we don't need form of any type.
        image = ArtPiecePhoto()
        image.filename = ''
        image.image=file
        image.piece = ArtPiece.objects.get(pk=id)
        image.save()
        print 'File saving done'
    
        #getting url for photo deletion
        file_delete_url = '/delete/'
    
        #getting file url here
        file_url = '/'
    
        #getting thumbnail url using sorl-thumbnail
        im = get_thumbnail(image, "80x80", quality=50)
        thumb_url = im.url
    
        #generating json response array
        result = []
        result.append({"name":filename, 
                       "size":file_size,
                       "id":image.pk,
                       "url":file_url, 
                       "thumbnail_url":thumb_url,
                       "delete_url":file_delete_url+str(image.pk)+'/', 
                       "delete_type":"POST",})
        response_data = json.dumps(result)
        return HttpResponse(response_data, content_type='application/json')
    else:
        return HttpResponse("", content_type='application/json')